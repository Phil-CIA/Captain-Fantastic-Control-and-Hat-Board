from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from io import TextIOWrapper
from pathlib import Path
from typing import Iterable

import openpyxl

from build_cross_reference_sheet import CrossReferenceRow, dedupe_rows, read_quote_rows, safe_int
from jlc_to_digikey_bom import exclude_from_bom, normalize_text, parse_quantity


SUPPORTED_SUFFIXES = {".csv", ".xls", ".xlsx"}


@dataclass(frozen=True)
class SourceScanResult:
    path: Path
    source_kind: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build one reusable master cross-reference sheet from a directory tree of JLC BOM CSVs and LCSC quote workbooks."
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help="A BOM file or a directory tree containing project BOM CSVs and LCSC quote workbooks.",
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        help="Output CSV path. Defaults to master_cross_reference.csv in the input directory.",
    )
    return parser.parse_args()


def find_input_files(input_path: Path) -> list[SourceScanResult]:
    if input_path.is_file():
        result = classify_source(input_path)
        return [result] if result is not None else []

    results: list[SourceScanResult] = []
    for path in sorted(input_path.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_SUFFIXES:
            continue
        result = classify_source(path)
        if result is not None:
            results.append(result)
    return results


def classify_source(path: Path) -> SourceScanResult | None:
    lower_name = path.name.lower()
    if path.suffix.lower() == ".csv" and "digikey" in lower_name:
        return None
    if path.suffix.lower() in {".xls", ".xlsx"}:
        if "cross_reference" in lower_name:
            return None
        return SourceScanResult(path=path, source_kind="workbook")
    if path.suffix.lower() == ".csv":
        return SourceScanResult(path=path, source_kind="bom_csv")
    return None


def read_csv_bom_rows(input_csv: Path) -> list[CrossReferenceRow]:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with input_csv.open("r", newline="", encoding=encoding) as handle:
                reader = csv.DictReader(handle)
                if reader.fieldnames is None:
                    return []

                rows: list[CrossReferenceRow] = []
                for row in reader:
                    if exclude_from_bom(row):
                        continue

                    lcsc_pn = normalize_text(row.get("LCSC", "") or row.get("LCSC#", ""))
                    manufacturer = normalize_text(row.get("Manufacturer", "") or row.get("MANUFACTURER", ""))
                    manufacturer_pn = normalize_text(
                        row.get("Manufacturer_Part_Number", "") or row.get("Part", "") or row.get("Manufacturer_PN", "")
                    )
                    value = normalize_text(row.get("Value", "") or row.get("Description", ""))
                    footprint = normalize_text(row.get("Footprint", "") or row.get("Package", ""))

                    if not any((lcsc_pn, manufacturer, manufacturer_pn, value)):
                        continue

                    rows.append(
                        CrossReferenceRow(
                            lcsc_pn=lcsc_pn,
                            manufacturer=manufacturer,
                            manufacturer_pn=manufacturer_pn,
                            description=value,
                            package=footprint,
                            quantity=parse_quantity(row),
                            stock_status="",
                            availability="",
                            matched_status="From_Project_BOM",
                            product_link="",
                            source_file=str(input_csv),
                        )
                    )
                return rows
        except UnicodeDecodeError:
            continue

    raise UnicodeDecodeError("batch-csv", b"", 0, 1, f"Could not decode CSV file {input_csv}")


def find_header_row_xlsx(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, list[str]]:
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [normalize_text(cell) for cell in row]
        if "Matched status" in values and "Product Link" in values:
            return row_index, values
    raise ValueError("Could not find LCSC quotation header row.")


def read_xlsx_quote_rows(input_xlsx: Path) -> list[CrossReferenceRow]:
    workbook = openpyxl.load_workbook(input_xlsx, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header_row_index, headers = find_header_row_xlsx(sheet)
    rows: list[CrossReferenceRow] = []

    for row_index, raw_row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_index <= header_row_index:
            continue
        values = [normalize_text(cell) for cell in raw_row]
        if not any(values):
            continue
        row = dict(zip(headers, values))
        matched_status = normalize_text(row.get("Matched status", ""))
        manufacturer_pn = normalize_text(row.get("Mrf#", ""))
        lcsc_pn = normalize_text(row.get("LCSC#", ""))

        if manufacturer_pn.startswith("-("):
            continue
        if not any((lcsc_pn, manufacturer_pn, row.get("Description", ""))):
            continue

        rows.append(
            CrossReferenceRow(
                lcsc_pn=lcsc_pn,
                manufacturer=normalize_text(row.get("Mfr.", "")),
                manufacturer_pn=manufacturer_pn,
                description=normalize_text(row.get("Description", "")),
                package=normalize_text(row.get("Package", "")),
                quantity=safe_int(normalize_text(row.get("Quantity", "0"))),
                stock_status=normalize_text(row.get("Stock Status", "")),
                availability=normalize_text(row.get("Availability", "")),
                matched_status=matched_status,
                product_link=normalize_text(row.get("Product Link", "")),
                source_file=str(input_xlsx),
            )
        )

    workbook.close()
    return rows


def read_source_rows(source: SourceScanResult) -> list[CrossReferenceRow]:
    if source.path.suffix.lower() == ".csv":
        return read_csv_bom_rows(source.path)
    if source.path.suffix.lower() == ".xls":
        return read_quote_rows(source.path)
    if source.path.suffix.lower() == ".xlsx":
        try:
            return read_xlsx_quote_rows(source.path)
        except ValueError:
            return []
    return []


def merge_source_files(rows: list[CrossReferenceRow]) -> list[CrossReferenceRow]:
    merged: dict[tuple[str, str, str], CrossReferenceRow] = {}
    sources_by_key: dict[tuple[str, str, str], set[str]] = {}

    for row in rows:
        key = (row.lcsc_pn, row.manufacturer, row.manufacturer_pn)
        existing = merged.get(key)
        sources_by_key.setdefault(key, set()).add(row.source_file)
        if existing is None:
            merged[key] = row
            continue

        replacement = existing
        if row.quantity > existing.quantity:
            replacement = row
        elif not existing.product_link and row.product_link:
            replacement = row
        elif existing.matched_status == "From_Project_BOM" and row.matched_status != "From_Project_BOM":
            replacement = row
        merged[key] = replacement

    merged_rows: list[CrossReferenceRow] = []
    for key, row in merged.items():
        source_files = " | ".join(sorted(sources_by_key[key]))
        merged_rows.append(
            CrossReferenceRow(
                lcsc_pn=row.lcsc_pn,
                manufacturer=row.manufacturer,
                manufacturer_pn=row.manufacturer_pn,
                description=row.description,
                package=row.package,
                quantity=row.quantity,
                stock_status=row.stock_status,
                availability=row.availability,
                matched_status=row.matched_status,
                product_link=row.product_link,
                source_file=source_files,
            )
        )

    return dedupe_rows(merged_rows)


def write_cross_reference(output_csv: Path, rows: list[CrossReferenceRow]) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "LCSC_PN",
                "Manufacturer",
                "Manufacturer_PN",
                "Description",
                "Package",
                "Matched_Status",
                "Stock_Status",
                "Availability",
                "Example_Qty",
                "DigiKey_PN",
                "Mouser_PN",
                "Onshore_Preferred_PN",
                "Onshore_Preferred_Supplier",
                "Preferred_Status",
                "Notes",
                "Product_Link",
                "Source_File",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "LCSC_PN": row.lcsc_pn,
                    "Manufacturer": row.manufacturer,
                    "Manufacturer_PN": row.manufacturer_pn,
                    "Description": row.description,
                    "Package": row.package,
                    "Matched_Status": row.matched_status,
                    "Stock_Status": row.stock_status,
                    "Availability": row.availability,
                    "Example_Qty": row.quantity,
                    "DigiKey_PN": "",
                    "Mouser_PN": "",
                    "Onshore_Preferred_PN": "",
                    "Onshore_Preferred_Supplier": "",
                    "Preferred_Status": "",
                    "Notes": "",
                    "Product_Link": row.product_link,
                    "Source_File": row.source_file,
                }
            )


def main() -> int:
    args = parse_args()
    sources = find_input_files(args.input_path)
    if not sources:
        raise SystemExit(f"No supported BOM or workbook files found under {args.input_path}")

    default_output_root = args.input_path if args.input_path.is_dir() else args.input_path.parent
    output_csv = args.output_csv or (default_output_root / "master_cross_reference.csv")

    all_rows: list[CrossReferenceRow] = []
    scanned_counts = {"bom_csv": 0, "workbook": 0}
    for source in sources:
        scanned_counts[source.source_kind] += 1
        all_rows.extend(read_source_rows(source))

    merged_rows = merge_source_files(all_rows)
    write_cross_reference(output_csv, merged_rows)

    print(f"Scanned files: {len(sources)}")
    print(f"  BOM CSV files: {scanned_counts['bom_csv']}")
    print(f"  Workbook files: {scanned_counts['workbook']}")
    print(f"Raw rows read: {len(all_rows)}")
    print(f"Master cross-reference rows: {len(merged_rows)}")
    print(f"Master cross-reference CSV: {output_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())